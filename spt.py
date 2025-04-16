# Shortest Paths Tracker 2022.1 [0430] | Eslam Abdullah

app_name = 'Shortest Paths Tracker 2022.1 | Eslam Abdullah\n'
file_name = ''
run_start = 0
run_end = 0
disconnected_pairs = 0


def run_duration():
    import datetime
    global file_name, run_start, run_end
    duration = str(datetime.timedelta(seconds=round(run_end - run_start)))
    msg_log(f'Elapsed Time: {duration}\n')


def msg_log(msg_text):
    global file_name
    with open(f'{file_name}_log.txt', 'a') as f:
        f.write(msg_text)


def create_log():
    import datetime
    global file_name
    now = datetime.datetime.now().strftime("%d-%m-%y %H:%M:%S")
    with open(f'{file_name}_log.txt', 'w') as f:
        f.write(f'{app_name}\nFile Name: {file_name}.xlsx\nLast Run: {now}\n\nWarning and Error Log:\n')


def write_paths(fs, origin_dijkstra, tail_index, sp_ws):
    print(f'Writing shortest paths for origin node {fs[0][tail_index]}...')
    from openpyxl import load_workbook
    global file_name
    wb = load_workbook(f'{file_name}.xlsx')
    ws = wb[sp_ws]
    main_text = ''
    nodes_names = fs[0]
    nodes_no = len(fs[0])
    origin_name = origin_dijkstra[0]
    global disconnected_pairs
    for index in range(nodes_no):
        od_text, cost_text, path_text, line_text = '', '', '', ''
        destination_name = nodes_names[index]
        if origin_name != destination_name:
            od_text += f'({origin_name}-{destination_name})'
            msg_1 = f'[Warning!] {od_text} pair is disconnected.'
            cost_text += str(origin_dijkstra[1][index])
            head_name = destination_name
            tail_name = origin_dijkstra[2][index]
            if tail_name not in nodes_names:
                print(msg_1)
                msg_log(msg_1 + '\n')
                disconnected_pairs += 1
                continue
            else:
                path_text += f'{str(head_name)[-1::-1]}-{str(tail_name)[-1::-1]}-'
                while_broken = False
                while tail_name != origin_name:
                    head_name = tail_name
                    if head_name not in nodes_names:
                        while_broken = True
                        break
                    else:
                        tail_index = nodes_names.index(head_name)
                        tail_name = origin_dijkstra[2][tail_index]
                        path_text += f'{str(tail_name)[-1::-1]}-'
                if while_broken:
                    print(msg_1)
                    msg_log(msg_1 + '\n')
                    disconnected_pairs += 1
                    continue
        if od_text != '':
            line_text += f'{od_text}\t{cost_text}\t'
            path_text = path_text[-2::-1]
            line_text += path_text
            main_text += line_text + '\n'
            ws.append([od_text, eval(cost_text), path_text])
    wb.save(f'{file_name}.xlsx')
    return main_text


def arc_cost(fs, tail_node, head_node):
    node_index = fs[0].index(tail_node)
    arc_index = fs[1][node_index] - 1
    step = 0
    cost_index = -1
    while cost_index < 0:
        if fs[4][arc_index + step] == head_node:
            cost_index = arc_index + step
        else:
            step += 1
    cij = fs[5][cost_index]
    return cij


def order_costs(fs_nodes, dijkstra_costs, star):
    nodes_no = len(fs_nodes)
    sorted_star = []
    sub_nodes = []
    sub_costs = []
    for index in range(nodes_no):
        if fs_nodes[index] in star:
            sub_nodes.append(fs_nodes[index])
            sub_costs.append(dijkstra_costs[index])
    sub_nodes_no = len(sub_nodes)
    while len(sub_nodes) > 0:
        for index in range(sub_nodes_no):
            if sub_costs[index] == min(sub_costs):
                sorted_star.append(sub_nodes.pop(index))
                sub_costs.pop(index)
                break
    return sorted_star


def update_cost(fs, dijkstra, update_index, origin_index):
    origin_node = fs[0][origin_index]
    new_cost = 0
    previous_node = dijkstra[1][update_index]
    head_node = fs[0][update_index]
    new_cost += arc_cost(fs, previous_node, head_node)
    while previous_node != origin_node:
        head_node = previous_node
        update_index = fs[0].index(head_node)
        previous_node = dijkstra[1][update_index]
        new_cost += arc_cost(fs, previous_node, head_node)
    return new_cost


def bellman_check(fs, dijkstra_costs, node_i, node_j):
    ci_index = fs[0].index(node_i)
    cj_index = fs[0].index(node_j)
    cj_ci = dijkstra_costs[cj_index] - dijkstra_costs[ci_index]
    cij = arc_cost(fs, node_i, node_j)
    if cj_ci <= cij:
        return True
    else:
        return False


def find_star(fs, tail_name):
    heads_names = []
    pointer_index = fs[0].index(tail_name)
    pointer_value = fs[1][pointer_index]
    if pointer_value <= len(fs[2]):
        arc_index = fs[1][pointer_index] - 1
        if pointer_index < (len(fs[0]) - 1):
            heads_no = fs[1][pointer_index + 1] - fs[1][pointer_index]
        else:
            heads_no = fs[2][-1] - fs[1][pointer_index] + 1
        step = 0
        for index in range(heads_no):
            heads_names.append(fs[4][arc_index + step])
            step += 1
    return heads_names


def iter_dijkstra(fs, dijkstra, origin_index):
    print(f'Analyzing shortest paths for origin node {fs[0][origin_index]}...')
    nodes_no = len(fs[0])
    visited_nodes = []
    while len(dijkstra[2]) > 0:
        current_node = dijkstra[2][0]
        visited_nodes.append(current_node)
        current_star = find_star(fs, current_node)
        dijkstra[2].pop(0)
        for head_node in current_star:
            if not bellman_check(fs, dijkstra[0], current_node, head_node):
                update_index = fs[0].index(head_node)
                dijkstra[1][update_index] = current_node
                new_cost = update_cost(fs, dijkstra, update_index, origin_index)
                dijkstra[0][update_index] = new_cost
        for head_node in current_star:
            if head_node not in dijkstra[2] and head_node not in visited_nodes:
                dijkstra[2].append(head_node)
        new_list = order_costs(fs[0], dijkstra[0], dijkstra[2])
        dijkstra[2] = new_list
    fs_dijkstra = [dijkstra[0], dijkstra[1]]
    for index in range(nodes_no):
        if fs[0][index] not in visited_nodes:
            fs_dijkstra[1][index] = 0
    return fs_dijkstra


def init_dijkstra(fs, dijkstra, origin_index):
    nodes_no = len(fs[0])
    origin_node = fs[0][origin_index]
    inf_cost = sum(fs[5]) + 1
    dijkstra[2].clear()
    dijkstra[2].append(origin_node)
    for index in range(nodes_no):
        dijkstra[1][index] = fs[0][origin_index]
        if index == origin_index:
            dijkstra[0][index] = 0
        else:
            dijkstra[0][index] = inf_cost
    return dijkstra


def process_paths(fs, sp_ws):
    print("Initializing Dijkstra's shortest path algorithm...")
    print_text = '\n(O-D)\tCost\tPath\n'
    nodes_no = len(fs[0])
    dijkstra_costs = [0 for node in range(nodes_no)]
    dijkstra_origins = [0 for node in range(nodes_no)]
    dijkstra_links = []
    dijkstra = [dijkstra_costs, dijkstra_origins, dijkstra_links]
    for origin_index in range(nodes_no):
        initialization = init_dijkstra(fs, dijkstra, origin_index)
        iter_results = iter_dijkstra(fs, initialization, origin_index)
        results = [fs[0][origin_index], iter_results[0], iter_results[1]]
        paths_text = write_paths(fs, results, origin_index, sp_ws)
        print_text += paths_text
    print("Printing shortest paths results...")
    print(print_text)
    global disconnected_pairs
    all_pairs = nodes_no**2 - nodes_no
    percent_disconnected = round((disconnected_pairs / all_pairs) * 100)
    if percent_disconnected > 0:
        msg_1 = f'[Warning!] {percent_disconnected}% of node pairs are disconnected.\n'
        print(msg_1)
        msg_log(msg_1)
    print('Run was successful.')
    msg_log('...\n\n')
    import time
    global run_end
    run_end = time.perf_counter()
    run_duration()
    import os
    import keyboard
    global file_name
    try:
        os.startfile(f'{file_name}.xlsx')
        print('Press any key to close.')
        while True:
            if keyboard.is_pressed(keyboard.read_key()):
                raise SystemExit
    except Exception:
        print('Press any key to close.')
        while True:
            if keyboard.is_pressed(keyboard.read_key()):
                raise SystemExit


def write_fs(fs, fs_ws, sp_ws):
    print("Writing forward star...")
    global file_name
    from openpyxl import load_workbook
    wb = load_workbook(f'{file_name}.xlsx')
    ws = wb[fs_ws]
    nodes_no = len(fs[0])
    arcs_no = len(fs[2])
    if arcs_no > nodes_no:
        for index in range(nodes_no):
            ws.append([fs[0][index], fs[1][index], None, fs[2][index], fs[3][index], fs[4][index], fs[5][index]])
        for index in range(nodes_no, arcs_no):
            ws.append([None, None, None, fs[2][index], fs[3][index], fs[4][index], fs[5][index]])
    if nodes_no > arcs_no:
        for index in range(arcs_no):
            ws.append([fs[0][index], fs[1][index], None, fs[2][index], fs[3][index], fs[4][index], fs[5][index]])
        for index in range(arcs_no, nodes_no):
            ws.append([fs[0][index], fs[1][index]])
    if nodes_no == arcs_no:
        for index in range(arcs_no):
            ws.append([fs[0][index], fs[1][index], None, fs[2][index], fs[3][index], fs[4][index], fs[5][index]])
    wb.save(f'{file_name}.xlsx')
    process_paths(fs, sp_ws)


def create_ws(fs):
    global file_name
    fs_ws = 'Forward Star'
    sp_ws = 'Shortest Paths'
    fs_title = ['Node', 'Pointer', None, 'Arc', 'Tail', 'Head', 'Cost']
    sp_title = ['(O-D)', 'Cost', 'Path']
    print(f"Creating '{fs_ws}' and '{sp_ws}' worksheets...")
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    try:
        wb = load_workbook(f'{file_name}.xlsx')
        current_sheets = wb.sheetnames
        fs_ws = f'{str(wb.active.title)} {fs_ws}'
        sp_ws = f'{str(wb.active.title)} {sp_ws}'
        try:
            if fs_ws in current_sheets:
                wb.remove(wb[fs_ws])
                wb.create_sheet(fs_ws)
                ws = wb[fs_ws]
                ws.append(fs_title)
                ws['A1'].font = ws['B1'].font = ws['D1'].font = ws['E1'].font = ws['F1'].font = ws['G1'].font = Font(
                    bold=True)
                wb.save(f'{file_name}.xlsx')
            else:
                wb.create_sheet(fs_ws)
                ws = wb[fs_ws]
                ws.append(fs_title)
                ws['A1'].font = ws['B1'].font = ws['D1'].font = ws['E1'].font = ws['F1'].font = ws['G1'].font = Font(
                    bold=True)
                wb.save(f'{file_name}.xlsx')
            if sp_ws in current_sheets:
                wb.remove(wb[sp_ws])
                wb.create_sheet(sp_ws)
                ws = wb[sp_ws]
                ws.append(sp_title)
                ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)
                wb.save(f'{file_name}.xlsx')
            else:
                wb.create_sheet(sp_ws)
                ws = wb[sp_ws]
                ws.append(sp_title)
                ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)
                wb.save(f'{file_name}.xlsx')
        except Exception as e:
            msg_1 = f'Creating Worksheet Error! {e}'
            msg_2 = f"Close '{file_name}.xlsx' workbook.\n"
            print(msg_1), print(msg_2)
            msg_log(msg_1 + '\n' + msg_2)
            print('Run failed.')
            import keyboard
            print('Press any key to close.')
            while True:
                if keyboard.is_pressed(keyboard.read_key()):
                    raise SystemExit
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = fs_ws
        ws.append(fs_title)
        ws['A1'].font = ws['B1'].font = ws['D1'].font = ws['E1'].font = ws['F1'].font = ws['G1'].font = Font(
            bold=True)
        wb.create_sheet(sp_ws)
        ws = wb[sp_ws]
        ws.append(sp_title)
        ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)
        wb.save(f'{file_name}.xlsx')
    write_fs(fs, fs_ws, sp_ws)


def combine_fs(fs_pointer, fs_arcs):
    fs = [fs_pointer[0].copy(), fs_pointer[1].copy(), fs_arcs[0].copy(), fs_arcs[1].copy(), fs_arcs[2].copy(),
          fs_arcs[3].copy()]
    create_ws(fs)


def create_fs_pointer(input_fs_arcs):
    print('Creating forward star...')
    arcs_no = len(input_fs_arcs[0])
    tail_names = list(dict.fromkeys(input_fs_arcs[0]))
    tail_names.sort()
    head_names = list(dict.fromkeys(input_fs_arcs[1]))
    head_names.sort()
    arcs_order = [0 for arc in range(arcs_no)]
    for arc_index in range(arcs_no):
        arcs_order[arc_index] = arc_index + 1
    old_tail, old_head, old_cost = input_fs_arcs[0].copy(), input_fs_arcs[1].copy(), input_fs_arcs[2].copy()
    new_tail, new_head, new_cost = [], [], []
    while len(old_tail) > 0:
        tail_min = min(old_tail)
        for arc_index in range(arcs_no):
            if old_tail[arc_index] == tail_min:
                new_tail.append(old_tail.pop(arc_index))
                new_head.append(old_head.pop(arc_index))
                new_cost.append(old_cost.pop(arc_index))
                break
    newer_tail, newer_head, newer_cost = [], [], []
    while len(new_tail) > 0:
        temp_tail = new_tail[0]
        temp_head = []
        for arc_index in range(len(new_tail)):
            if new_tail[arc_index] == temp_tail:
                temp_head.append(new_head[arc_index])
            else:
                break
        temp_head.sort()
        for arc_index in range(len(temp_head)):
            temp_index = new_head.index(temp_head[arc_index])
            newer_tail.append(new_tail.pop(temp_index))
            newer_head.append(new_head.pop(temp_index))
            newer_cost.append(new_cost.pop(temp_index))
        temp_head.clear()
    sorted_fs_arcs = [arcs_order, newer_tail, newer_head, newer_cost]
    nodes_names = tail_names.copy()
    for node in head_names:
        if node not in nodes_names:
            nodes_names.append(node)
    nodes_no = len(nodes_names)
    print('Checking number of nodes...')
    msg_1 = "[Graph Limit Error!] Number of nodes can't exceed 1000."
    msg_2 = 'Reduce graph size by eliminating irrelevant nodes and arcs.\n'
    if nodes_no > 1000:
        print(msg_1), print(msg_2)
        msg_log(msg_1 + '\n' + msg_2)
        init_program()
    nodes_pointer = [0 for arc in range(nodes_no)]
    for node in range(nodes_no):
        if nodes_names[node] in newer_tail:
            nodes_pointer[node] = newer_tail.index(nodes_names[node]) + 1
        else:
            nodes_pointer[node] = arcs_no + 1
    sorted_fs_pointer = [nodes_names, nodes_pointer]
    combine_fs(sorted_fs_pointer, sorted_fs_arcs)


def manual_fs_arcs(arcs_no):
    arcs_tail, arcs_head = [0 for arc in range(arcs_no)], [0 for arc in range(arcs_no)]
    arcs_cost = [-1 for arc in range(arcs_no)]
    for index in range(arcs_no):
        print(f'Arc {index + 1} : ')
        insert_arc = True
        while insert_arc:
            insert_arc = False
            while arcs_tail[index] < 1:
                temp = input('Tail > ')
                arcs_tail[index] = int(temp) if temp.isnumeric() else 0
            while arcs_head[index] < 1:
                temp = input('Head > ')
                if temp.isnumeric():
                    if int(temp) != arcs_tail[index]:
                        arcs_head[index] = int(temp)
            for moving_index in range(index):
                if arcs_tail[moving_index] == arcs_tail[index] and arcs_head[moving_index] == arcs_head[index]:
                    print(f'Arc ({arcs_tail[index]}-{arcs_head[index]}) already exists!')
                    arcs_tail[index], arcs_head[index] = 0, 0
                    insert_arc = True
                    break
        while arcs_cost[index] < 0:
            arcs_cost[index] = input('Cost > ').strip()
            try:
                arcs_cost[index] = eval(arcs_cost[index]) if eval(arcs_cost[index]) >= 0 else -1
            except Exception:
                arcs_cost[index] = -1
        print(f'C({arcs_tail[index]}-{arcs_head[index]}) = {arcs_cost[index]}\n')
    import time
    global run_start
    run_start = time.perf_counter()
    fs_arcs = [arcs_tail, arcs_head, arcs_cost]
    create_log()
    create_fs_pointer(fs_arcs)


def info_check(fs_arcs):
    info_ok = [True, True, True, True, True]
    for tail in fs_arcs[0]:
        msg_1 = '[Graph Info Error!] [Col. A] Tail nodes must have positive integer names.'
        if type(tail) != int:
            info_ok[0] = False
            print(msg_1)
            msg_log(msg_1 + '\n')
        elif tail < 1:
            info_ok[0] = False
            print(msg_1)
            msg_log(msg_1 + '\n')
        if not info_ok[0]:
            break
    for head in fs_arcs[1]:
        msg_1 = '[Graph Info Error!] [Col. B] Head nodes must have positive integer names.'
        if type(head) != int:
            info_ok[1] = False
            print(msg_1)
            msg_log(msg_1 + '\n')
        elif head < 1:
            info_ok[1] = False
            print(msg_1)
            msg_log(msg_1 + '\n')
        if not info_ok[1]:
            break
    for cost in fs_arcs[2]:
        msg_1 = '[Graph Info Error!] [Col. C] Costs can only assume non-negative decimals.'
        if type(cost) != int and type(cost) != float:
            info_ok[2] = False
            print(msg_1)
            msg_log(msg_1 + '\n')
        elif cost < 0:
            info_ok[2] = False
            print(msg_1)
            msg_log(msg_1 + '\n')
        if not info_ok[2]:
            break
    arcs_no = len(fs_arcs[0])
    arcs_nonsense = []
    arcs_repeated = []
    for i in range(arcs_no):
        fixed_arc = [fs_arcs[0][i], fs_arcs[1][i]]
        if fixed_arc[0] == fixed_arc[1]:
            nonsense = f'({fixed_arc[0]}-{fixed_arc[1]})'
            if nonsense not in arcs_nonsense:
                arcs_nonsense.append(nonsense)
                info_ok[3] = False
        for j in range(arcs_no):
            moving_arc = [fs_arcs[0][j], fs_arcs[1][j]]
            if i != j and fixed_arc == moving_arc:
                repeated = f'({moving_arc[0]}-{moving_arc[1]})'
                if repeated not in arcs_repeated:
                    arcs_repeated.append(repeated)
                    info_ok[4] = False
    for nonsense in arcs_nonsense:
        msg_1 = f'[Graph Info Error!] [Col. A:B] Arc {nonsense} has no sense.'
        print(msg_1)
        msg_log(msg_1 + '\n')
    for repeated in arcs_repeated:
        msg_2 = f'[Graph Info Error!] [Col. A:B] Arc {repeated} is assigned more than once.'
        print(msg_2)
        msg_log(msg_2 + '\n')
    all_ok = True
    for check in info_ok:
        if check is False:
            all_ok = False
    return all_ok


def read_graph():
    print('Checking file...')
    global file_name
    from openpyxl import load_workbook
    try:
        wb = load_workbook(f'{file_name}.xlsx')
        ws = wb.active
        columns = ws.columns
    except Exception as e:
        msg_1 = f'[File Not Found!] {e}'
        msg_2 = 'Check file name and directory of the application.\n'
        print(msg_1), print(msg_2)
        file_name = ''
        init_program()

    print('Importing graph info...')
    arcs_tail, arcs_head, arcs_cost, arcs_nway = [], [], [], []
    nway_values = [None, 1, 2]
    msg_3 = '[Info Import Error!] One or more columns seems to be entirely blank.'
    msg_4 = 'Check active worksheet and graph information.\n'
    create_log()
    try:
        for i in range(4):
            col = [cell.value for cell in next(columns)]
            if i == 0:
                arcs_tail = col[1:]
            if i == 1:
                arcs_head = col[1:]
            if i == 2:
                arcs_cost = col[1:]
            if i == 3:
                arcs_nway = col[1:]
    except Exception:
        print(msg_3), print(msg_4)
        msg_log(msg_3 + '\n' + msg_4)
        init_program()

    print('Checking graph info...')
    try:
        for index in range(len(arcs_nway)):
            if arcs_nway[index] in nway_values:
                if arcs_nway[index] is None:
                    arcs_nway[index] = 1
                if arcs_nway[index] == 2:
                    arcs_tail.append(arcs_head[index])
                    arcs_head.append(arcs_tail[index])
                    arcs_cost.append(arcs_cost[index])
            else:
                msg_1 = '[Graph Info Error!] [Col. D] Number of directions must be 1 or 2 (blank = 1).'
                print(msg_1), print(msg_4)
                msg_log(msg_1 + '\n' + msg_4)
                init_program()
        fs_arcs = [arcs_tail, arcs_head, arcs_cost]
        if len(arcs_tail) == len(arcs_head) == len(arcs_cost) == sum(arcs_nway) > 0:
            if info_check(fs_arcs):
                create_fs_pointer(fs_arcs)
            else:
                print(msg_4)
                msg_log(msg_4)
                init_program()
        else:
            print(msg_3), print(msg_4)
            msg_log(msg_3 + '\n' + msg_4)
            init_program()
    except Exception as e:
        msg_1 = f'[Unexpected Error!] {e}\n'
        print(msg_1)
        msg_log(msg_1)
        init_program()


def remove_char(text, chars):
    for char in chars:
        text = text.replace(char, '')
    text = text.strip()
    return text


def init_program():
    chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|', '.xlsx']
    global file_name
    file_name = ''
    while file_name == '':
        file_name = input('File Name [.xlsx] > ')
        file_name = remove_char(file_name, chars)
    import keyboard
    print('Press [R] to read or [W] to write.')
    while True:
        if keyboard.is_pressed('r'):
            import time
            global run_start
            run_start = time.perf_counter()
            print()
            print('Input Mode: Read.')
            read_graph()
            break
        if keyboard.is_pressed('w'):
            print()
            print('Input Mode: Write.')
            arcs_no = 0
            while arcs_no < 1:
                arcs_no = input('Number of Arcs > ')
                arcs_no = int(arcs_no) if arcs_no.isnumeric() else 0
            print()
            manual_fs_arcs(arcs_no)
            break


print(app_name)
init_program()
